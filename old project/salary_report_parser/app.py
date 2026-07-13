from salary_report_parser.worker import Worker, Job
from openpyxl import load_workbook
import logging
from logs.logging_module import logger, generate_handler
import os

log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")

os.makedirs(log_path, exist_ok=True)
logger.addHandler(generate_handler(os.path.join(log_path, "salary_report_parser_all.log"), logging.DEBUG))
logger.addHandler(generate_handler(os.path.join(log_path, "salary_report_parser_info.log"), logging.INFO))
logger.setLevel(logging.DEBUG)
logger.info(f"Logger enabled")


def safe_stoi_convertion(s: str | None) -> int | None:
    if s is None:
        return s
    try:
        return int(s)
    except (ValueError, TypeError) as e:
        logger.error(f"Fail during stoi convertion {e} from string: {s}")
        return None


# A slobby amoeba
def parse_excel_report(path: str) -> (dict[str, Worker], str):
    path = os.path.abspath(path)

    logger.info(f"Trying to parse: {path}")
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    last_sheet_name = sheet_names[-1]
    sheet = workbook[last_sheet_name]
    workers_dict = {}
    date = sheet.cell(row=3, column=3).value.strftime("%d-%m-%Y")

    for row in sheet.iter_rows(min_row=8, values_only=True):
        if row[0] is not None and row[1] is not None:
            # RE DO
            work_type = row[4]
            mark = row[5]
            tonns = row[6]
            runs = row[7]
            hectars = row[8]
            hours = row[9]
            salary_for_day = row[12]
            job = Job(
                work_type=work_type,
                mark=round(safe_stoi_convertion(mark), 2) if safe_stoi_convertion(mark) else None,
                tonns=round(safe_stoi_convertion(tonns), 2)if safe_stoi_convertion(tonns) else None,
                runs=round(safe_stoi_convertion(runs), 2)if safe_stoi_convertion(runs) else None,
                hectars=round(safe_stoi_convertion(hectars), 2) if safe_stoi_convertion(hectars) else None,
                hours=round(safe_stoi_convertion(hours), 2) if safe_stoi_convertion(hours) else None ,
                salary_for_day=round(safe_stoi_convertion(salary_for_day), 2) if safe_stoi_convertion(salary_for_day) else None)

            worker = Worker(
                unique_id=row[1],
                name=row[2],
                machine_type=row[3],
                # commentary=row[4],
                hours_worked_sum=round(safe_stoi_convertion(row[10]), 2) if safe_stoi_convertion(row[10]) else None,
                days_worked=row[11],
                salary_for_month=round(safe_stoi_convertion(row[13]), 2) if safe_stoi_convertion(row[13]) else None,
                repair_days_count=row[14],
                absence_reason=row[15],
                jobs=[job]
            )
            if workers_dict.get(worker.unique_id) is None:
                workers_dict[worker.unique_id] = worker
            else:
                workers_dict[worker.unique_id].add_jobs_from(worker)
        if row[1] is None:
            logger.info("Special data is skipped")
    logger.info(f"Parsed successfully")
    workbook.close()
    return workers_dict, date

# test

# workers, date = parse_excel_report("../unparsed_reports/report.xlsx")
# logger.info(f"Printing result messages:\n")
# for worker in workers:
#     logger.info(f"{worker.generate_message(date)}")
# logger.info(f"Finished printing result messages:\n")
#
# logger.info(f"Moving report to the ../parsed_reports/{date}report.xlsx folder:\n")
# shutil.move("../unparsed_reports/report.xlsx", f"../parsed_reports/{date}report.xlsx")
# logger.info(f"Moved successfully\n")
